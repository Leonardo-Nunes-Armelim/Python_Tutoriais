from openpyxl import load_workbook
import datetime
import matplotlib.pyplot as plt

# Carregando uma planilha Excel para a variavel "work_book"
work_book = load_workbook(filename='Excel de exemplo.xlsx')

work_sheet = work_book['Produtos']

columns = len(work_sheet[1]) # 6 colunas
rows = len(work_sheet['A'])  # 12 linhas

# Criando gráfico a partir de Excel
# Coordenada dos Dados
x = []
y = []
soma = 0

for row in range(2, rows):
    x.append(int(datetime.datetime.strptime(str(work_sheet.cell(row, 1).value.date()), "%Y-%m-%d").timetuple().tm_year))
    y.append(work_sheet.cell(row, 3).value)
    soma += int(work_sheet.cell(row, 2).value) * int(work_sheet.cell(row, 3).value)

print('Soma', soma)

figura, eixo = plt.subplots()
eixo.scatter(x, y, s=200)
eixo.grid(True)
plt.show()
