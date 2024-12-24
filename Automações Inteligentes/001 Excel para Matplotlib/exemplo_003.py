from openpyxl import load_workbook
import time
import datetime
import matplotlib.pyplot as plt

# Carregando uma planilha Excel para a variavel "work_book"
work_book = load_workbook(filename='Excel de exemplo.xlsx')

work_sheet = work_book['Produtos']

columns = len(work_sheet[1]) # 6 colunas
rows = len(work_sheet['A'])  # 12 linhas

# Criando gráfico a partir de Excel
# Coordenada dos Dados
quantidade = []
items = []
soma = 0

for row in range(2, rows):
    if work_sheet.cell(row, 6).value not in ['iPhone', 'iPod']:
        soma += int(work_sheet.cell(row, 2).value * int(work_sheet.cell(row, 3).value))

for row in range(2, rows):
    if work_sheet.cell(row, 6).value not in ['iPhone', 'iPod']:
        quantidade.append((work_sheet.cell(row, 2).value * int(work_sheet.cell(row, 3).value) / soma))
        items.append(work_sheet.cell(row, 6).value)

print('Soma', soma)

figura, eixo = plt.subplots()
eixo.pie(quantidade, labels=items, autopct='%1.1f%%')
eixo.grid(True)
plt.show()
