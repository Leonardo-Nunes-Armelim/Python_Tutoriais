from openpyxl import Workbook

# Cria um objeto de um Excel
work_book = Workbook()

# Pega a aba da planilha ativa
work_sheet = work_book.active

# Indice das linhas e colunas da planilha
#   |____A____|____B____|____C____|
# 1 |__Nome___|__Idade__|_Altura__|
# 2 |_Leo_____|___25____|__1,75___|
# 3 |_Cesar___|___18____|__1,65___|
# 4 |_Davi____|___50____|__1,70___|
# 5 |_Carol___|___26____|__1,70___|

data = [[ 'Nome', 'Idade', 'Altura'],
        [  'Leo',      25,     1.75],
        ['Cesar',      18,     1.65],
        [ 'Davi',      50,     1.70],
        ['Carol',      26,     1.70]]

for row in data:
    work_sheet.append(row)

# Modo 1: Pela chave da celula
modo_1 = work_sheet['A1'].value
print('>>> modo_1 A1:', modo_1)

# Modo 2: Pela posição da celula
modo_2 = work_sheet.cell(row=2, column=1).value
print('>>> modo_2 A2:', modo_2)

# Modo 3: Por um loop
for column in range(1, 4):
    for row in range(1, 6):
        print(f'row {row} e column {column}:', work_sheet.cell(row=row, column=column).value)

# Modo 4: Range nas celulas
cell_range = work_sheet['A1':'C2']
print(cell_range)
for column in range(len(cell_range[0])):
    for row in range(len(cell_range)):
        print(f'>>> col_range row {row} column {column}', cell_range[row][column].value)

# Modo 5: Range de coluna
column_C = work_sheet['C']
print(column_C)
for row in range(len(column_C)):
    print(f'>>> column_C row {row}', column_C[row].value)

# Modo 6: Range de colunas
column_range = work_sheet['B:C']
print(column_range)
for column in range(len(column_range)):
    for row in range(len(column_range[0])):
        print(f'>>> column_range row {row} column {column}', column_range[column][row].value)

# Modo 7: Range de Linha
row_2 = work_sheet[2]
print(row_2)
for column in range(len(row_2)):
    print(f'>>> column column {column}', row_2[column].value)

# Modo 8: Range de linhas
row_range = work_sheet[1:2]
print(row_range)
for column in range(len(row_range)):
    for row in range(len(row_range[0])):
        print(f'>>> row_range row {row} column {column}', row_range[column][row].value)

# Modo 9: Range em formato de loop percorrendo as linhas
for row in work_sheet.iter_rows(min_row=1, max_col=3, max_row=2):
    print(row)
    for cell in row:
        print('>>> iter_rows:', cell.value)

# Modo 10: Range em formato de loop percorrendo as colunas
for column in work_sheet.iter_cols(min_row=1, max_col=3, max_row=2):
    print(column)
    for cell in column:
        print('>>> iter_cols:', cell.value)

# Modo 11: Interação por linhas
rows = tuple(work_sheet.rows)
print(rows)
for row in range(len(rows)):
    for column in range(len(rows[0])):
        print('>>>', rows[row][column].value)

# Modo 12: Interação por colunas
columns = tuple(work_sheet.columns)
print(columns)
for column in range(len(columns)):
    for row in range(len(columns[0])):
        print('>>>', columns[column][row].value)

# Modo 13: Loop por todos os valores da planilha
print(tuple(work_sheet.values))
for row in work_sheet.values:
    for value in row:
        print(value)

# Modo 14: Interação por linhas
for row in work_sheet.iter_rows(min_row=1, max_col=3, max_row=3, values_only=True):
    print(row)
    for value in row:
        print(value)

# Modo 15: Interação por colunas
for column in work_sheet.iter_cols(min_row=1, max_col=3, max_row=3, values_only=True):
    print(column)
    for value in column:
        print(value)
