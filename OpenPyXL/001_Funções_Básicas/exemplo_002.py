from openpyxl import Workbook

# Cria um objeto de um Excel
work_book = Workbook()

# Pega a aba da planilha ativa (Seleciona a aba de indice 0)
work_sheet = work_book.active

# Alterando nome da planilha
work_sheet.title = 'Plan Inicial'

# Insere uma planilha no final (padrão)
work_sheet_1 = work_book.create_sheet('Nova Plan 1')
# Insere uma planilha na posição inicial
work_sheet_2 = work_book.create_sheet('Nova Plan 2', 0)
# Insere uma planilha na penultima posição
work_sheet_3 = work_book.create_sheet('Nova Plan 3', -1)

# Deletando planilha
work_book.remove(work_book['Plan Inicial'])

# Movendo planilha da terceira posição para a primeira posição
work_book.move_sheet(work_book['Nova Plan 1'], offset=-2)

# Vendo todos os nomes das planilhas
print('work_book.sheetnames', work_book.sheetnames)

# Salvando o objeto em um arquivos excel
work_book.save("sample.xlsx")
