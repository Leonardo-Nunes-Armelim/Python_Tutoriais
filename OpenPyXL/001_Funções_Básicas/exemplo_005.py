from openpyxl import load_workbook

# Carregando uma planilha Excel para a variavel "work_book"
work_book = load_workbook(filename='planilha.xlsx')

# Planilha #######################################

work_sheet = work_book['Plan1']

# Escrever #######################################

# Modo 1 de inserir valores no Excel (pela chave da celula)
work_sheet['A8'] = 'Curta o vídeo'

# Modo 2 de inserir valores no Excel (pela posição da celula)
work_sheet.cell(row=9, column=1, value='Se inscreva no canal')

# Modo 3 de inserir valores no Excel (pelo objeto da celula)
cell = work_sheet['A10']
cell.value = 'Compartilhe o vídeo'

# Modo 4 de inserir valores no Excel (atravéz de uma lista)
data = [[ 'Nome', 'Idade', 'Altura'],
        [  'Leo',      25,     1.75],
        ['Cesar',      18,     1.65],
        [ 'Davi',      50,     1.70],
        ['Carol',      26,     1.70]]

for row in data:
    work_sheet.append(row)

# Ler ############################################

# Modo 1: Pela chave da celula
modo_1 = work_sheet['A1'].value
print('>>> modo_1 A1:', modo_1)

# Modo 2: Pela posição da celula
modo_2 = work_sheet.cell(row=2, column=1).value
print('>>> modo_2 A2:', modo_2)

# Salvando o objeto em um arquivos excel
work_book.save("sample_2.xlsx")
